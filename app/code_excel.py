from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def summary(S0, K, T, r, q, sigma, optionType, style,sMin, sMax,sSteps, tSteps, method,BC,omega, tol, price, bsPrice,yellow, gray, lime):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    input_labels = [
        'Asset Price', 'Strike Price(K)', 'Maturity (Years)', 'Interest Rate(p.a)','Div yield(p.a)', 'Volatility (p.a)', 'Option Type', 'Exercise Style',
        'Min. Asset Price', 'Max. Asset Price', 'Asset steps', 'Time steps']
    input_values = [S0, K, T, r, q, sigma, optionType, style, sMin, sMax, sSteps, tSteps]

    output_labels = [f"{style} {optionType} Premium", 'Black-Merton Analytic']
    output_values = [price, bsPrice]
    # Header
    ws['B2'] = 'Inputs'
    ws['B2'].fill = yellow
    ws['C2'].fill = gray
    
    # Inputs
    for i, (label, value) in enumerate(zip(input_labels, input_values), start=3):
        ws[f'B{i}'] = label
        ws[f'B{i}'].fill = gray
        ws[f'C{i}'] = value
        ws[f'C{i}'].fill = yellow

    # Outputs
    ws['B17'] = 'Outputs'
    ws['B17'].fill = yellow
    ws['C17'].fill = gray
    for i, (label, value) in enumerate(zip(output_labels, output_values), start=18):
        ws[f'B{i}'] = label
        ws[f'B{i}'].fill = gray
        ws[f'C{i}'] = value
        ws[f'C{i}'].fill = yellow

    # Additional info for non-European options
    if style != 'European':
        ws['E13'] = f'Omega: {omega}'
        ws['E13'].font = Font(italic=True, size=10)
        ws['E14'] = f'Tolerance: 1e-{tol}'
        ws['E14'].font = Font(italic=True, size=10)

    ws.column_dimensions['B'].width = 19
    return wb

def bermudanDates(wb, dates):
    ws = wb['FDM Grid']
    ws['A3'] = 'Can Exercise on'
    ws['A3'].font = Font(bold=True)
    ws['A3'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    for i, date in enumerate(dates):
        cell = f'A{4 + i}'
        ws[cell] = date
        ws[cell].fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")

    ws.column_dimensions['A'].width = 13
    return wb

def print_excel(style, sMin, sMax, sSteps, tSteps, method, BC, omega, tol,optionType, A, B, grid, is_exercised,
                   S0, K, T, r, q, sigma, price, bsPrice, exercise_dates,alphaGamma):
    blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    yellow = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    gray = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
    lime = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red = Font(color="FF0000")

    # Create workbook and summary sheet
    wb = summary(S0, K, T, r, q, sigma, optionType,style,sMin, sMax, sSteps, tSteps, method, BC,omega, tol, price, bsPrice,yellow, gray, lime)
    ws = wb.create_sheet('FDM Grid')
    startR, startC = 4, 3
    if style == 'Bermudan':
        startC += 1
        wb = bermudanDates(wb, exercise_dates)
        ws = wb['FDM Grid']

    # Title
    col = number_to_letter(startC)
    ws[f'{col}{startR - 2}'] = f"Pricing the {style} {optionType} using the {method} FDM ({BC}-BC)"
    ws[f'{col}{startR - 2}'].font = Font(size=10)
    ws[f'A1'], ws[f'A1'].font = "Red: Exercised Option", Font(color="FF0000")

    # Axis labels
    _col = number_to_letter(startC - 1)
    ws[f'{_col}{startR - 1}'] = 'Time/Asset'
    ws[f'{_col}{startR - 1}'].font = Font(size=9)
    ws[f'{_col}{startR}'], ws[f'{_col}{startR}'].fill = 0, gray
    ws[f'{_col}{startR + sSteps}'], ws[f'{_col}{startR + sSteps}'].fill =sSteps, gray

    # Coefficient matrix labels alphaGamma
    coefName = f"Matrix A ({method})"
    coefRow = startR + sSteps + 3
    coefColB = startC + sSteps + 2
    alphaGammaLabel = ["alpha(1)", f"gamma({sSteps-1})"]
    if BC == "Neumann":
        alphaGamma = ["",""]
        alphaGammaLabel = ["",""]
    if method == "Crank":
        coefName = "Matrix A (Implicit)"
        ws[f'{number_to_letter(coefColB)}{coefRow - 1}'] = "Matrix B (Explicit)"
        ws[f'{number_to_letter(coefColB+sSteps-1)}{coefRow}'] = alphaGammaLabel[0]
        ws[f'{number_to_letter(coefColB+sSteps-1)}{coefRow+1}'] = alphaGammaLabel[1]
        ws[f'{number_to_letter(coefColB+sSteps)}{coefRow}'] = -alphaGamma[0] if len(alphaGammaLabel[0])>1 else ""
        ws[f'{number_to_letter(coefColB+sSteps)}{coefRow+1}'] = -alphaGamma[1] if len(alphaGammaLabel[1])>1 else ""
        
    ws[f'{number_to_letter(startC)}{coefRow - 1}'] = coefName
    ws[f'{number_to_letter(coefColB-3)}{coefRow}'] = alphaGammaLabel[0]
    ws[f'{number_to_letter(coefColB-3)}{coefRow+1}'] = alphaGammaLabel[1]
    ws[f'{number_to_letter(coefColB-2)}{coefRow}'] = alphaGamma[0]
    ws[f'{number_to_letter(coefColB-2)}{coefRow+1}'] = alphaGamma[1]
        
    # Grid and coefficients
    coefCol_index = 0
    for t in range(tSteps + 1):
        colLetter = number_to_letter(startC + t)
        coef_col_B = number_to_letter(coefColB + t)
        # Time header
        ws[f'{colLetter}{startR - 1}'] = t
        ws[f'{colLetter}{startR - 1}'].fill = gray
        ws[f'{colLetter}{startR - 2}'].fill = yellow
        # First asset row
        ws[f'{colLetter}{startR}'] = grid[0, t]
        ws[f'{colLetter}{startR}'].fill = blue

        for s in range(1, sSteps):
            row = startR + s
            if t == 0:
                ws[f'{_col}{row}'] = s
                ws[f'{_col}{row}'].fill = gray
                
            ws[f'{colLetter}{row}'] = grid[s, t]
            ws[f'{colLetter}{row}'].fill = yellow
                
            if is_exercised[s, t]:
                ws[f'{colLetter}{row}'].font = red

            if coefCol_index < sSteps - 1:
                coefRow_index = coefRow + s - 1
                ws[f'{colLetter}{coefRow_index}'] = A[s - 1, coefCol_index]
                ws[f'{colLetter}{coefRow_index}'].fill = lime
                if method == "Crank":
                    ws[f'{coef_col_B}{coefRow_index}'] = B[s - 1, coefCol_index]
                    ws[f'{coef_col_B}{coefRow_index}'].fill = gray

        coefCol_index += 1
        # Last asset row
        ws[f'{colLetter}{startR + sSteps}'] = grid[-1, t]
        ws[f'{colLetter}{startR + sSteps}'].fill = blue

    # Fill remaining coefficients if sSteps > tSteps
    for j in range(coefCol_index, sSteps-1):
        colLetter = number_to_letter(j + startC)
        coef_col_B = number_to_letter(coefColB + j)
        for i in range(sSteps-1):
            row = coefRow + i
            ws[f'{colLetter}{row}'] = A[i, j]
            ws[f'{colLetter}{row}'].fill = lime
            if method == "Crank":
                ws[f'{coef_col_B}{row}'] = B[i, j]
                ws[f'{coef_col_B}{row}'].fill = gray
    return wb

def get_workbook_as_bytes(wb):
    with BytesIO() as buffer:
        wb.save(buffer)
        return buffer.getvalue()

def number_to_letter(n):
    n = max(n, 1)
    letters = []
    while n:
        n, remainder = divmod(n - 1, 26)
        letters.append(chr(65 + remainder))
    return ''.join(reversed(letters))
